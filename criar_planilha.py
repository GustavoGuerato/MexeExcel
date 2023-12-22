import openpyxl
from random import uniform

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 0)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 120), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(5, 16):
    planilha2.cell(linha, 1).value = f'Gustavo{linha}{round(uniform(10, 120), 2)}'
    planilha2.cell(linha, 1).value = f'Gabriel{linha}{round(uniform(10, 120), 2)}'
    planilha2.cell(linha, 1).value = f'Silvia{linha}{round(uniform(10, 120), 2)}'

planilha.save('planilha_nova.xlsx')
