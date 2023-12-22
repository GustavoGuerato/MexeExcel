import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nomes_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']


for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 120), 2)
    planilha1.cell(linha,3).value = preco

for linha in planilha1:
    if linha[0].value is not None:
        print(linha[0].value, end=' ')
    if linha[1].value is not None:
        print(linha[1].value, end=' ')
    if linha[2].value is not None:
        print(linha[2].value)

pedidos.save('nova_planilha.xlsx')
